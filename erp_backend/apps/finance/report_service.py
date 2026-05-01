"""
Report Generation Service
===========================
Dynamic query builder that executes report definitions and exports
results to PDF, Excel, CSV, or JSON.
"""
import csv
import io
import json
import logging
import os
from datetime import datetime
from decimal import Decimal

logger = logging.getLogger(__name__)

# Model registry — maps data_source names to Django models
MODEL_REGISTRY = {}


def _build_registry():
    """Lazily build model registry on first use."""
    global MODEL_REGISTRY
    if MODEL_REGISTRY:
        return

    from erp.connector_registry import connector

    # POS
    POSOrder = connector.require('pos.orders.get_model', org_id=0, source='finance.report')
    POSOrderLine = connector.require('pos.order_lines.get_model', org_id=0, source='finance.report')
    if POSOrder is not None:
        MODEL_REGISTRY['POSOrder'] = POSOrder
        MODEL_REGISTRY['Order'] = POSOrder  # Alias
    if POSOrderLine is not None:
        MODEL_REGISTRY['OrderLine'] = POSOrderLine

    # Finance - Invoices
    try:
        from apps.finance.invoice_models import Invoice, InvoiceLine
        MODEL_REGISTRY['Invoice'] = Invoice
        MODEL_REGISTRY['InvoiceLine'] = InvoiceLine
    except ImportError: pass

    # Finance - Models (same module, direct import OK)
    try:
        from apps.finance.models import Voucher, Asset, DeferredExpense, FinancialAccount, ChartOfAccount
        MODEL_REGISTRY['Voucher'] = Voucher
        MODEL_REGISTRY['Asset'] = Asset
        MODEL_REGISTRY['DeferredExpense'] = DeferredExpense
        MODEL_REGISTRY['FinancialAccount'] = FinancialAccount
        MODEL_REGISTRY['ChartOfAccount'] = ChartOfAccount
    except ImportError: pass

    # Finance - Payments (same module, direct import OK)
    try:
        from apps.finance.payment_models import Payment
        MODEL_REGISTRY['Payment'] = Payment
    except ImportError: pass

    # Inventory
    Product = connector.require('inventory.products.get_model', org_id=0, source='finance.report')
    Category = connector.require('inventory.categories.get_model', org_id=0, source='finance.report')
    Warehouse = connector.require('inventory.warehouses.get_model', org_id=0, source='finance.report')
    if Product is not None:
        MODEL_REGISTRY['Product'] = Product
    if Category is not None:
        MODEL_REGISTRY['Category'] = Category
    if Warehouse is not None:
        MODEL_REGISTRY['Warehouse'] = Warehouse
    # Unit + StockAlert have no connector capability yet — fall back to direct
    # import inside try/except (Pattern D for legacy models with no connector entry).
    try:
        from apps.inventory.models import Unit  # noqa: E402  (Pattern D: no capability yet)
        MODEL_REGISTRY['Unit'] = Unit
    except ImportError: pass
    try:
        from apps.inventory.alert_models import StockAlert  # noqa: E402  (Pattern D: no capability yet)
        MODEL_REGISTRY['StockAlert'] = StockAlert
    except ImportError: pass

    # CRM
    Contact = connector.require('crm.contacts.get_model', org_id=0, source='finance.report')
    if Contact is not None:
        MODEL_REGISTRY['Contact'] = Contact

    # HR — Employee has a connector capability; Attendance/Leave do not yet.
    Employee = connector.require('hr.employees.get_model', org_id=0, source='finance.report')
    if Employee is not None:
        MODEL_REGISTRY['Employee'] = Employee
    try:
        from apps.hr.models import Attendance, Leave  # noqa: E402  (Pattern D: no capability yet)
        MODEL_REGISTRY['Attendance'] = Attendance
        MODEL_REGISTRY['Leave'] = Leave
    except ImportError: pass

    # Integrations — no connector_service.py for integrations yet
    try:
        from apps.integrations.models import ExternalOrderMapping, ExternalProductMapping  # noqa: E402  (Pattern D: no capability yet)
        MODEL_REGISTRY['ExternalOrderMapping'] = ExternalOrderMapping
        MODEL_REGISTRY['ExternalProductMapping'] = ExternalProductMapping
    except ImportError: pass


def _resolve_field(obj, field_path):
    """Resolve a dotted field path like 'contact.name' or 'invoice.balance_due'."""
    parts = field_path.split('.')
    current = obj
    for part in parts:
        if current is None:
            return None
        try:
            current = getattr(current, part, None)
            # If it's a callable (property/method with no args), call it
            if callable(current) and not hasattr(current, 'pk'):
                try:
                    current = current()
                except TypeError:
                    pass
        except Exception:
            return None
    if isinstance(current, Decimal):
        return float(current)
    elif isinstance(current, datetime):
        return current.isoformat()
    elif hasattr(current, 'pk'):
        return str(current)
    return current


# Filter operator mapping
FILTER_OPS = {
    'eq': '',           # exact match
    'ne': '',           # handled with exclude
    'gt': '__gt',
    'gte': '__gte',
    'lt': '__lt',
    'lte': '__lte',
    'contains': '__icontains',
    'icontains': '__icontains',
    'startswith': '__istartswith',
    'istartswith': '__istartswith',
    'in': '__in',
    'isnull': '__isnull',
    'range': '__range',
}


class ReportService:
    """
    Dynamic report generation engine.

    Usage:
        service = ReportService(organization_id)
        data = service.execute(report_definition)
        file_path = service.export(report_definition, data, format='EXCEL')
    """

    def __init__(self, organization_id):
        self.organization_id = organization_id
        _build_registry()

    def execute(self, report_def):
        """
        Execute a report definition and return the result data.

        Args:
            report_def: ReportDefinition instance

        Returns:
            dict with columns, rows, aggregations, row_count
        """
        model_class = MODEL_REGISTRY.get(report_def.data_source)
        if not model_class:
            return {'error': f"Unknown data source: {report_def.data_source}"}

        # Build base queryset (tenant-filtered)
        qs = model_class.objects.filter(organization_id=self.organization_id)

        # Apply filters
        for f in (report_def.filters or []):
            field = f.get('field')
            op = f.get('op', 'eq')
            value = f.get('value')

            if op == 'ne':
                qs = qs.exclude(**{field: value})
            elif op == 'in':
                qs = qs.filter(**{f"{field}__in": value if isinstance(value, list) else [value]})
            elif op == 'isnull':
                qs = qs.filter(**{f"{field}__isnull": value})
            else:
                lookup = FILTER_OPS.get(op, '')
                qs = qs.filter(**{f"{field}{lookup}": value})

        # Apply ordering
        if report_def.order_by:
            qs = qs.order_by(*report_def.order_by)

        # Apply limit
        qs = qs[:report_def.limit]

        # Extract column fields
        column_fields = [c['field'] for c in (report_def.columns or [])]
        if not column_fields:
            # Default: all fields
            column_fields = [f.name for f in model_class._meta.get_fields() if hasattr(f, 'column')]

        # Build rows
        rows = []
        for obj in qs:
            row = {}
            for field in column_fields:
                try:
                    row[field] = _resolve_field(obj, field)
                except Exception:
                    row[field] = None
            rows.append(row)

        # Compute aggregations
        agg_results = {}
        if report_def.aggregations:
            from django.db.models import Sum, Avg, Count, Min, Max
            agg_funcs = {'sum': Sum, 'avg': Avg, 'count': Count, 'min': Min, 'max': Max}
            agg_kwargs = {}
            for agg in report_def.aggregations:
                func = agg_funcs.get(agg.get('func', 'sum'))
                if func:
                    label = agg.get('label', agg['field'])
                    agg_kwargs[label] = func(agg['field'])
            if agg_kwargs:
                base_qs = model_class.objects.filter(organization_id=self.organization_id)
                for f in (report_def.filters or []):
                    field = f.get('field')
                    op = f.get('op', 'eq')
                    value = f.get('value')
                    if op == 'ne':
                        base_qs = base_qs.exclude(**{field: value})
                    else:
                        lookup = FILTER_OPS.get(op, '')
                        base_qs = base_qs.filter(**{f"{field}{lookup}": value})
                agg_results = base_qs.aggregate(**agg_kwargs)
                # Convert Decimals
                agg_results = {k: float(v) if isinstance(v, Decimal) else v for k, v in agg_results.items()}

        return {
            'columns': report_def.columns or [{'field': f, 'label': f.replace('_', ' ').title()} for f in column_fields],
            'rows': rows,
            'row_count': len(rows),
            'aggregations': agg_results,
        }

    def export_csv(self, report_def, data):
        """Export report data to CSV string."""
        output = io.StringIO()
        columns = data.get('columns', [])
        writer = csv.DictWriter(output, fieldnames=[c['field'] for c in columns])
        writer.writeheader()
        for row in data.get('rows', []):
            writer.writerow(row)
        return output.getvalue()

    def export_json(self, report_def, data):
        """Export report data to JSON string."""
        return json.dumps(data, indent=2, default=str)

    def export_excel(self, report_def, data, output_path):
        """
        Export report data to Excel file.
        Requires openpyxl.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.error("openpyxl not installed — cannot export Excel")
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = report_def.name[:31]

        columns = data.get('columns', [])

        # Header row
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='4F46E5')
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col.get('label', col['field']))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = col.get('width', 18)

        # Data rows
        for row_idx, row in enumerate(data.get('rows', []), 2):
            for col_idx, col in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col['field']))

        # Aggregation row
        if data.get('aggregations'):
            row_idx = len(data.get('rows', [])) + 3
            ws.cell(row=row_idx, column=1, value='TOTALS').font = Font(bold=True)
            for label, value in data['aggregations'].items():
                ws.cell(row=row_idx + 1, column=1, value=label).font = Font(bold=True)
                ws.cell(row=row_idx + 1, column=2, value=value)
                row_idx += 1

        wb.save(output_path)
        logger.info(f"[Report] Excel exported to {output_path}")
        return output_path

    def run_and_export(self, report_def, export_format=None, output_dir=None):
        """
        Execute report and export to file.

        Returns:
            dict with file_path, row_count, format
        """
        from django.utils import timezone

        fmt = export_format or report_def.default_export_format
        data = self.execute(report_def)

        if 'error' in data:
            return data

        if output_dir is None:
            output_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'media', 'reports')
        os.makedirs(output_dir, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_name = report_def.name.replace(' ', '_')[:50]

        if fmt == 'CSV':
            file_path = os.path.join(output_dir, f"{safe_name}_{timestamp}.csv")
            content = self.export_csv(report_def, data)
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)
        elif fmt == 'JSON':
            file_path = os.path.join(output_dir, f"{safe_name}_{timestamp}.json")
            content = self.export_json(report_def, data)
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)
        elif fmt == 'EXCEL':
            file_path = os.path.join(output_dir, f"{safe_name}_{timestamp}.xlsx")
            self.export_excel(report_def, data, file_path)
        else:
            file_path = os.path.join(output_dir, f"{safe_name}_{timestamp}.json")
            content = self.export_json(report_def, data)
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(content)

        # Update report execution timestamp
        report_def.last_run_at = timezone.now()
        report_def.save(update_fields=['last_run_at'])

        return {
            'file_path': file_path,
            'row_count': data['row_count'],
            'format': fmt,
            'aggregations': data.get('aggregations', {}),
        }
